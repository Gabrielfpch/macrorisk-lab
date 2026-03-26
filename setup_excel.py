#!/usr/bin/env python3
"""
setup_excel.py
==============
Genera la estructura base del archivo Excel para el Sistema de Requerimientos.

Pontificia Universidad Catolica del Peru - INFOPUC
Excel Avanzado: Macros - Trabajo Grupal 2

Uso:
    pip install openpyxl
    python setup_excel.py

Resultado:
    Crea "Requerimientos_Mantenimiento.xlsx" listo para convertir a .xlsm
    e importar el codigo VBA desde la carpeta /VBA/
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ============================================================
# CONSTANTES DEL SISTEMA
# ============================================================

COLOR_AZUL_OSCURO = "1F4E79"    # Azul marino institucional
COLOR_AZUL_CLARO  = "BDD7EE"    # Azul claro para subtitulos
COLOR_AMARILLO    = "FFFFC8"    # Amarillo para instrucciones
COLOR_BEIGE       = "F0F0E4"    # Beige para impresion
COLOR_BLANCO      = "FFFFFF"
COLOR_NEGRO       = "000000"
COLOR_GRIS        = "D6D6D6"

# Encabezados de la base de datos (orden exacto requerido)
ENCABEZADOS_REGISTRO = [
    "N. REQ",
    "TIPO DE MANT.",
    "TIPO DE TRABAJO",
    "UBICACION",
    "FECHA",
    "JUSTIFICACION",
    "PROVEEDOR",
    "RUC",
    "PRODUCTO",
    "MARCA",
    "MODELO",
    "COD. ACT.",
    "REFERENCIA",
]

# Anchos de columna para la hoja Registro (en caracteres)
ANCHOS_REGISTRO = [10, 14, 16, 22, 12, 35, 30, 14, 22, 15, 15, 14, 20]

# Datos maestros del sistema
UBICACIONES = [
    "UCI",
    "UCI-COVID",
    "Emergencia Pediatrica",
    "Emergencia Adultos",
    "Sala de Operaciones",
    "Traumatologia",
    "Reumatologia",
    "Sala de Maquinas",
]

TIPOS_TRABAJO = ["Mecanico", "Electrico", "Electronico", "Inspeccion"]

PRODUCTOS = [
    "Ascensores",
    "Bombas de agua",
    "Bombas de succion",
    "Grupos electrogenos",
    "Equipos biomedicos",
]

# (RUC ficticios - uso academico)
PROVEEDORES = [
    ("Electrobombas Peruanas SA",       "20100458560", "Motores y equipos de succion"),
    ("COVIEM SA",                        "20501234567", "Grupos electrogenos"),
    ("Ascensores Schindler del Peru SA", "20264853390", "Ascensores"),
    ("HeathCare",                        "20378901234", "Equipos biomedicos"),
    ("Proveedor General",                "",            "General"),
]


# ============================================================
# ESTILOS REUTILIZABLES
# ============================================================

def estilo_encabezado():
    """Estilo para celdas de encabezado de tabla."""
    return {
        "font":      Font(bold=True, color=COLOR_BLANCO, size=11),
        "fill":      PatternFill("solid", fgColor=COLOR_AZUL_OSCURO),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border":    borde_fino(),
    }

def estilo_subtitulo():
    return {
        "font":      Font(size=12, italic=True),
        "fill":      PatternFill("solid", fgColor=COLOR_AZUL_CLARO),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }

def borde_fino():
    lado = Side(style="thin")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def borde_medio():
    lado = Side(style="medium")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def aplicar_estilo(cell, estilo: dict):
    for attr, val in estilo.items():
        setattr(cell, attr, val)


# ============================================================
# HOJA: Formulario (pagina de bienvenida)
# ============================================================

def crear_hoja_formulario(wb: Workbook):
    ws = wb.active
    ws.title = "Formulario"

    # --- Titulo principal ---
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "SISTEMA DE REQUERIMIENTOS DE MANTENIMIENTO"
    c.font      = Font(bold=True, size=16, color=COLOR_BLANCO)
    c.fill      = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # --- Subtitulo ---
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value     = "Clinica Internacional San Francisco de Asis"
    c.font      = Font(size=13, italic=True)
    c.fill      = PatternFill("solid", fgColor=COLOR_AZUL_CLARO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    ws.row_dimensions[3].height = 15  # separador

    # --- Instruccion de uso ---
    ws.merge_cells("A4:G4")
    c = ws["A4"]
    c.value     = "Haga clic en el boton 'Abrir Formulario' para registrar un nuevo requerimiento"
    c.font      = Font(size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 22

    ws.row_dimensions[5].height = 12  # separador

    # --- Caja de instrucciones ---
    ws.merge_cells("A6:G13")
    c = ws["A6"]
    c.value = (
        "PASOS PARA AGREGAR EL BOTON DE ACCESO:\n\n"
        "  1. Vaya al menu: Insertar > Ilustraciones > Formas\n"
        "  2. Seleccione un Rectangulo redondeado\n"
        "  3. Dibuje el boton debajo de estas instrucciones\n"
        "  4. Escriba en el boton: 'Abrir Formulario de Requerimiento'\n"
        "  5. Clic derecho sobre el boton > 'Asignar macro'\n"
        "  6. Seleccione la macro: AbrirFormularioRequerimiento\n\n"
        "  (Recuerde que el archivo debe guardarse como .xlsm para habilitar macros)"
    )
    c.font      = Font(size=10)
    c.fill      = PatternFill("solid", fgColor=COLOR_AMARILLO)
    c.alignment = Alignment(horizontal="left", vertical="top",
                             wrap_text=True, indent=1)
    for r in range(6, 14):
        ws.row_dimensions[r].height = 18

    # Ajustar anchos de columna
    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


# ============================================================
# HOJA: Registro (base de datos)
# ============================================================

def crear_hoja_registro(wb: Workbook):
    ws = wb.create_sheet("Registro")

    estilo = estilo_encabezado()

    # --- Encabezados ---
    for col, nombre in enumerate(ENCABEZADOS_REGISTRO, start=1):
        c = ws.cell(row=1, column=col, value=nombre)
        c.font      = estilo["font"]
        c.fill      = estilo["fill"]
        c.alignment = estilo["alignment"]
        c.border    = estilo["border"]
        ws.column_dimensions[get_column_letter(col)].width = ANCHOS_REGISTRO[col - 1]

    ws.row_dimensions[1].height = 24

    # --- Fila de ejemplo (comentada en gris claro) ---
    ejemplo = [
        "MNTO1", "Preventivo", "Mecanico", "UCI",
        "01/04/2024", "Mantenimiento preventivo programado anual del ascensor principal.",
        "Ascensores Schindler del Peru SA", "20264853390",
        "Ascensores", "Schindler", "5300A", "ACT-001", "REF-2024-001",
    ]
    for col, val in enumerate(ejemplo, start=1):
        c = ws.cell(row=2, column=col, value=val)
        c.font      = Font(color="AAAAAA", italic=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Nota sobre la fila de ejemplo
    ws.row_dimensions[2].height = 20

    # Nota aclaratoria
    ws.merge_cells("A3:M3")
    c = ws["A3"]
    c.value = (
        "NOTA: La fila 2 es un ejemplo de referencia. "
        "Los registros reales se insertan desde el formulario VBA a partir de la fila 2 "
        "(borre esta fila de ejemplo antes de usar el sistema)."
    )
    c.font      = Font(size=8, italic=True, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 16

    # Congelar la fila de encabezados
    ws.freeze_panes = "A2"


# ============================================================
# HOJA: Impresion (placeholder - se genera por VBA)
# ============================================================

def crear_hoja_impresion(wb: Workbook):
    ws = wb.create_sheet("Impresion")

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = "Esta hoja se genera automaticamente al hacer clic en 'Imprimir Formato'"
    c.font      = Font(size=11, italic=True, color="888888")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value     = "No edite esta hoja manualmente."
    c.font      = Font(size=10, color="AAAAAA")
    c.alignment = Alignment(horizontal="center")

    for col_idx in range(1, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


# ============================================================
# HOJA: Datos (referencia de listas maestras)
# ============================================================

def crear_hoja_datos(wb: Workbook):
    """Hoja oculta con las listas maestras del sistema (referencia)."""
    ws = wb.create_sheet("Datos")

    encabezado_style = {
        "font":      Font(bold=True, color=COLOR_BLANCO),
        "fill":      PatternFill("solid", fgColor=COLOR_AZUL_OSCURO),
        "alignment": Alignment(horizontal="center"),
    }

    # --- Columna A: Ubicaciones ---
    c = ws["A1"]
    c.value = "UBICACIONES"
    for attr, val in encabezado_style.items():
        setattr(c, attr, val)
    for i, item in enumerate(UBICACIONES, start=2):
        ws.cell(row=i, column=1, value=item)
    ws.column_dimensions["A"].width = 24

    # --- Columna B: Tipos de Trabajo ---
    c = ws["B1"]
    c.value = "TIPOS DE TRABAJO"
    for attr, val in encabezado_style.items():
        setattr(c, attr, val)
    for i, item in enumerate(TIPOS_TRABAJO, start=2):
        ws.cell(row=i, column=2, value=item)
    ws.column_dimensions["B"].width = 18

    # --- Columna C: Productos/Equipos ---
    c = ws["C1"]
    c.value = "PRODUCTOS / EQUIPOS"
    for attr, val in encabezado_style.items():
        setattr(c, attr, val)
    for i, item in enumerate(PRODUCTOS, start=2):
        ws.cell(row=i, column=3, value=item)
    ws.column_dimensions["C"].width = 22

    # --- Columnas D-F: Proveedores ---
    for col, label in zip(["D", "E", "F"],
                          ["PROVEEDOR", "RUC", "AREA"]):
        c = ws[f"{col}1"]
        c.value = label
        for attr, val in encabezado_style.items():
            setattr(c, attr, val)
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 28

    for i, (nombre, ruc, area) in enumerate(PROVEEDORES, start=2):
        ws.cell(row=i, column=4, value=nombre)
        ws.cell(row=i, column=5, value=ruc)
        ws.cell(row=i, column=6, value=area)

    ws.sheet_state = "hidden"   # Ocultar esta hoja auxiliar


# ============================================================
# FUNCION PRINCIPAL
# ============================================================

def crear_workbook() -> Workbook:
    wb = Workbook()
    crear_hoja_formulario(wb)
    crear_hoja_registro(wb)
    crear_hoja_impresion(wb)
    crear_hoja_datos(wb)
    return wb


def main():
    nombre_archivo = "Requerimientos_Mantenimiento.xlsx"

    print("=" * 60)
    print("  PUCP - INFOPUC | Excel Avanzado: Macros - Trabajo 2")
    print("  Generando archivo Excel base...")
    print("=" * 60)

    wb = crear_workbook()
    wb.save(nombre_archivo)

    print(f"\n  Archivo creado: {nombre_archivo}")
    print("\n" + "=" * 60)
    print("  PASOS SIGUIENTES:")
    print("=" * 60)
    print()
    print("  1. Abra el archivo en Microsoft Excel")
    print()
    print("  2. Guarde como libro habilitado para macros:")
    print("     Archivo > Guardar como > Tipo: 'Libro de Excel")
    print("     habilitado para macros (*.xlsm)'")
    print()
    print("  3. Abra el Editor de Visual Basic: Alt + F11")
    print()
    print("  4. Importe Module1.bas:")
    print("     Archivo > Importar archivo > VBA/Module1.bas")
    print()
    print("  5. Inserte el UserForm:")
    print("     Insertar > UserForm")
    print("     Cambie el nombre a 'frmRequerimiento' (ventana Propiedades)")
    print()
    print("  6. Disenio del UserForm:")
    print("     Consulte INSTRUCCIONES.md para la lista completa de controles")
    print("     y sus nombres exactos (Name property)")
    print()
    print("  7. Pegue el codigo del UserForm:")
    print("     Doble clic en el form > copie todo VBA/frmRequerimiento.vba")
    print()
    print("  8. Configure la hoja Formulario:")
    print("     Alt+F8 > ConfiguracionCompleta > Ejecutar")
    print()
    print("  9. Guarde el archivo .xlsm y pruebe haciendo clic en el boton")
    print()
    print("  Listo! Ver INSTRUCCIONES.md para mas detalles.")
    print("=" * 60)


if __name__ == "__main__":
    main()
